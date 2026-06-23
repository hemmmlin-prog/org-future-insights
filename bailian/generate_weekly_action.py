"""
bailian/generate_weekly_action.py — 每周五自动生成 HRBP 行动计划 Excel

由 launchd 在周五 18:00 自动调用，也可手动执行：
    python3 -m bailian.generate_weekly_action          # 本周
    python3 -m bailian.generate_weekly_action 2026-06-13  # 指定周五日期

输出：组织演变/weekly-actions/YYYY-WNN-行动计划.xlsx
"""
from __future__ import annotations

import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from bailian.client import chat
from bailian.generate_daily import WORKSPACE, RAW_DIR, REPORT_DIR

WEEKLY_DIR = WORKSPACE / "weekly-actions"
WEEKLY_DIR.mkdir(parents=True, exist_ok=True)

SYSTEM_PROMPT = """你是一位资深 HR 顾问，服务对象是速卖通 HRBP Leader（支持供给团队和供应链团队）。

供给团队岗位：商家运营、类目招商 BD、选品策略、商家服务
供应链团队岗位：物流路径规划、需求预测、仓储调度、跨境物流协调

请基于本周的多日 HR 洞察日报，输出一份结构化的周度行动计划。

输出要求（严格 JSON 格式）：
{
  "week_summary": "本周核心趋势一句话",
  "top_signals": ["信号1", "信号2", "信号3"],
  "actions": [
    {
      "priority": "high|medium|low",
      "timeframe": "本周|下周|本月|本季",
      "action": "具体行动项（跟谁做什么得到什么）",
      "owner": "自己|HRG团队|业务Leader|自己+HRG",
      "deliverable": "可验证的产出物",
      "signal_source": "关联的信号来源"
    }
  ],
  "positions_impact": [
    {
      "category": "放大型|转型型|压缩型",
      "position": "岗位名称",
      "team": "供给|供应链",
      "ai_impact": "AI影响描述",
      "suggested_action": "建议动作"
    }
  ]
}

规则：
- actions 应有 6-9 项，覆盖本周(2-3项)/下周(2项)/本月(2项)/本季(1-2项)
- positions_impact 应有 4-6 项，覆盖放大型/转型型/压缩型
- 行动项禁止出现"关注""了解""思考"等非行动动词
- 必须具体到速卖通供给/供应链场景
- 所有 signal_source 必须引用输入中实际存在的来源"""


def collect_week_reports(friday: date) -> str:
    """收集 Mon-Fri 的 PM/auto 日报内容"""
    monday = friday - timedelta(days=4)
    reports = []

    for i in range(5):
        d = monday + timedelta(days=i)
        ds = d.strftime("%Y-%m-%d")

        # 优先 PM 版，没有则用 auto 版
        pm_path = REPORT_DIR / f"{ds}-pm.md"
        auto_path = REPORT_DIR / f"{ds}-auto.md"

        if pm_path.exists():
            content = pm_path.read_text()
            reports.append(f"=== {ds} (PM精读版) ===\n{content[:3000]}")
        elif auto_path.exists():
            content = auto_path.read_text()
            reports.append(f"=== {ds} (自动版) ===\n{content[:3000]}")

    if not reports:
        print("⚠️ 本周无任何日报，跳过生成")
        return ""

    print(f"📄 收集到 {len(reports)} 天日报")
    return "\n\n".join(reports)


def generate_plan(week_content: str, model: str = "qwen-max") -> dict:
    """调用百炼生成周度行动计划"""
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"以下是本周的 HR 洞察日报汇总，请生成周度行动计划：\n\n{week_content}"},
    ]

    print(f"🤖 调用 {model} 生成行动计划...")
    resp = chat(messages, model=model, temperature=0.4, response_format={"type": "json_object"})
    print(f"  ✅ 生成完成 | {resp.elapsed_sec:.1f}s | {resp.input_tokens}→{resp.output_tokens} tokens")

    return json.loads(resp.text)


def write_excel(plan: dict, friday: date) -> Path:
    """将行动计划写入 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    week_num = friday.isocalendar()[1]
    week_label = f"{friday.year}-W{week_num:02d}"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="2F5496")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')

    priority_map = {"high": "🔴 高", "medium": "🟠 中", "low": "🟡 低"}

    # ===== Sheet 1: 行动计划 =====
    ws = wb.active
    ws.title = "周度行动计划"

    ws.merge_cells('A1:F1')
    ws['A1'] = f'速卖通 HRBP Leader 周度行动计划（{week_label}）'
    ws['A1'].font = title_font

    ws.merge_cells('A2:F2')
    ws['A2'] = f'本周核心：{plan.get("week_summary", "")}'
    ws['A2'].font = Font(italic=True, size=11)

    # Top signals
    ws.merge_cells('A3:F3')
    signals = plan.get("top_signals", [])
    ws['A3'] = f'三大信号：① {signals[0] if len(signals) > 0 else ""} ② {signals[1] if len(signals) > 1 else ""} ③ {signals[2] if len(signals) > 2 else ""}'
    ws['A3'].font = Font(size=10, color="666666")

    # Headers
    headers = ['优先级', '时间窗', '行动项', '负责人', '产出物', '关联信号']
    col_widths = [10, 10, 45, 14, 28, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Action rows
    actions = plan.get("actions", [])
    for i, action in enumerate(actions):
        row = 6 + i
        values = [
            priority_map.get(action.get("priority", ""), action.get("priority", "")),
            action.get("timeframe", ""),
            action.get("action", ""),
            action.get("owner", ""),
            action.get("deliverable", ""),
            action.get("signal_source", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border
        ws.row_dimensions[row].height = 45

    # ===== Sheet 2: 岗位影响分档 =====
    ws2 = wb.create_sheet("岗位AI影响分档")
    ws2.merge_cells('A1:E1')
    ws2['A1'] = f'供给 & 供应链 AI 岗位影响分档（{week_label}）'
    ws2['A1'].font = title_font

    headers2 = ['分档', '岗位', '所属团队', 'AI 影响描述', '建议动作']
    col_widths2 = [14, 20, 10, 38, 28]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    row = 3
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    category_map = {"放大型": "⬆️ 放大型", "转型型": "➡️ 转型型", "压缩型": "⬇️ 压缩型"}

    positions = plan.get("positions_impact", [])
    for i, pos in enumerate(positions):
        row = 4 + i
        values = [
            category_map.get(pos.get("category", ""), pos.get("category", "")),
            pos.get("position", ""),
            pos.get("team", ""),
            pos.get("ai_impact", ""),
            pos.get("suggested_action", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border
        ws2.row_dimensions[row].height = 35

    # Save
    output_path = WEEKLY_DIR / f"{week_label}-行动计划.xlsx"
    wb.save(output_path)
    return output_path


def main():
    # 确定目标周五
    if len(sys.argv) > 1:
        friday = datetime.strptime(sys.argv[1], "%Y-%m-%d").date()
    else:
        today = date.today()
        # 找到本周五（如果今天是周五就用今天）
        days_until_friday = (4 - today.weekday()) % 7
        friday = today + timedelta(days=days_until_friday) if days_until_friday > 0 else today

    week_num = friday.isocalendar()[1]
    print(f"📅 生成周度行动计划: {friday} (W{week_num:02d})")

    # 1. 收集本周日报
    week_content = collect_week_reports(friday)
    if not week_content:
        sys.exit(0)

    # 2. 调百炼生成
    plan = generate_plan(week_content)

    # 3. 写 Excel
    output = write_excel(plan, friday)
    print(f"✅ Excel 已生成: {output}")
    print(f"   相对路径: {output.relative_to(WORKSPACE)}")


if __name__ == "__main__":
    main()
